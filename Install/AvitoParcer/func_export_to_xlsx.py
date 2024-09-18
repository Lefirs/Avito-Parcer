from openpyxl import load_workbook, Workbook



def export_to_xlsx(objects, namefile):
    wb = Workbook() # Инициализация файла xlsx

    row = 1

    for i in wb.sheetnames:
        wb.remove(wb[i]) # Удаление всех возможных листов

    SheetObjects = wb.create_sheet('Part1_0-50') # Создание нового листа

    for i in objects:
        # Добавление нового содержимого в лист
        print(i.get('Name'))
        SheetObjects[f'A{str(row)}'] = i.get('Name')
        SheetObjects[f'B{str(row)}'] = i.get('Address')
        SheetObjects[f'C{str(row)}'] = i.get('Link')
        SheetObjects[f'D{str(row)}'] = i.get('Price')
        row = row + 1

    wb.save(f'{namefile}.xlsx')